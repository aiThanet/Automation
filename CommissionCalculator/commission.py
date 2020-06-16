import pyodbc 
import openpyxl
import csv
import os
import sys
import tkinter
from tkinter import messagebox
from tkinter import filedialog
from openpyxl.styles import colors
from openpyxl.styles import Font, Color
from collections import defaultdict

root = tkinter.Tk()
extra_commission = {}

class EditorClass(object):
    UPDATE_PERIOD = 100 #ms
    editors = []
    updateId = None
    def __init__(self, master):
        self.__class__.editors.append(self)
        self.lineNumbers = ''
        # A frame to hold the three components of the widget.
        self.frame = tkinter.Frame(master, bd=1, relief=tkinter.SUNKEN)
        self.frame.pack(side=tkinter.TOP)
        # The widgets vertical scrollbar
        self.vScrollbar = tkinter.Scrollbar(self.frame, orient=tkinter.VERTICAL)
        self.vScrollbar.pack(fill='y', side=tkinter.RIGHT)
        # The Text widget holding the line numbers.
        self.lnText = tkinter.Text(self.frame,
                width = 4,height=15,
                highlightthickness = 0,
                takefocus = 0,
                bd = 0,
                background = 'lightgrey',
                foreground = 'magenta',
                state='disabled'
        )
        self.lnText.pack(side=tkinter.LEFT, fill='y')
        # The Main Text Widget
        self.text = tkinter.Text(self.frame,
                width=36,height=15,
                bd=0,
                undo=True,
                background = 'white'
        )
        self.text.pack(side=tkinter.LEFT, fill=tkinter.BOTH, expand=1)
        self.text.config(yscrollcommand=self.vScrollbar.set)
        self.vScrollbar.config(command=self.text.yview)
        if self.__class__.updateId is None:
            self.updateAllLineNumbers()
    def getLineNumbers(self):
        x = 0
        line = '0'
        col= ''
        ln = ''
        # assume each line is at least 6 pixels high
        step = 6
        nl = '\n'
        lineMask = '    %s\n'
        indexMask = '@0,%d'
        for i in range(0, self.text.winfo_height(), step):
            ll, cc = self.text.index( indexMask % i).split('.')
            if line == ll:
                if col != cc:
                    col = cc
                    ln += nl
            else:
                line, col = ll, cc
                ln += (lineMask % line)[-5:]
        return ln
    def updateLineNumbers(self):
        tt = self.lnText
        ln = self.getLineNumbers()
        if self.lineNumbers != ln:
            self.lineNumbers = ln
            tt.config(state='normal')
            tt.delete('1.0', tkinter.END)
            tt.insert('1.0', self.lineNumbers)
            tt.config(state='disabled')
    @classmethod
    def updateAllLineNumbers(cls):
        if len(cls.editors) < 1:
            cls.updateId = None
            return
        for ed in cls.editors:
            ed.updateLineNumbers()
        cls.updateId = ed.text.after(
            cls.UPDATE_PERIOD,
            cls.updateAllLineNumbers)

def scan_bill(filename='scan.xlsx'):
    scan_wb = openpyxl.load_workbook(filename=filename)
    scan_sheet = scan_wb['Sheet1']
    row_scan = 1

    bills = set()
    while True:
        if scan_sheet.cell(row=row_scan, column=1).value:
            bills.add(scan_sheet.cell(row=row_scan, column=1).value)
        else:
            break
        row_scan += 1
    return bills


def get_data(bills,server='127.0.0.1',database='database',username='uid',password = 'password',rate_engine=4,rate_part=6):
    global extra_commission
    output = {}
    config = 'DRIVER={SQL Server};' + f'SERVER={server};DATABASE={database};UID={username};PWD={password}'
    cnxn = pyodbc.connect(config)
    cursor = cnxn.cursor()

    for bill in bills:
        query_1 = '''
            SELECT *
            FROM SOInvHD INNER JOIN EMCust ON (SOInvHD.CustID=EMCust.CustID)
            WHERE DocuNo=?
            '''
        cursor.execute(query_1,bill)
        for row in cursor.fetchall():
            
            query_2 = '''
            SELECT *
            FROM SOInvDT INNER JOIN EMGood ON (SOInvDT.GoodID=EMGood.GoodID)
            WHERE SOInvID=?
            '''
            amt_total = 0
            cursor.execute(query_2, row.SOInvID)
            if row.CustName not in output:
                output[row.CustName] = defaultdict(float)
            for row2 in cursor.fetchall():
                amt_total += float(row2.GoodAmnt)
                
                if row2.GoodCode in extra_commission:
                    
                    output[row.CustName][bill + '_' + str(extra_commission[row2.GoodCode])] += float(row2.GoodAmnt)
                else :
                    if row2.GoodGroupID == 1000:
                        output[row.CustName][bill + '_' + str(rate_engine)] += float(row2.GoodAmnt)
                    elif row2.GoodGroupID == 1001:
                        output[row.CustName][bill + '_' + str(rate_part)] += float(row2.GoodAmnt)
   
    return output


def write_output(output, dest_filename='output.xlsx'):
    output_wb = openpyxl.Workbook()
    output_worksheet = output_wb.active

    col_name = ['ชื่อร้าน', 'เลขที่บิล(BL)','ยอดบิล','ธนาคาร','เลขที่เช็ค','เช็คผ่าน','ยอดเช็ค','เปอร์เซ็นต์','ยอดคอม']

    for i,col in enumerate(col_name):
        output_worksheet.cell(row=1, column=i+1).value = col

    curr_row = 2
    for customer in output:
        output_worksheet.cell(row=curr_row, column=1).value = customer
        total_amount = 0
        total_commission = 0
        for bill in sorted(output[customer]):
            if float(output[customer][bill]) > 0:
                bill_code, percent = bill.split("_")
                commission = round(output[customer][bill] * (float(percent)/100),2)

                output_worksheet.cell(row=curr_row, column=2).value = bill_code
                output_worksheet.cell(row=curr_row, column=3).value = output[customer][bill]
                output_worksheet.cell(row=curr_row, column=8).value = percent + "%"
                output_worksheet.cell(row=curr_row, column=9).value = commission

                total_amount += output[customer][bill]
                total_commission += commission

                curr_row += 1

        
        output_worksheet.cell(row=curr_row, column=2).value = 'รวม'
        output_worksheet.cell(row=curr_row, column=2).font = Font(bold=True)
        output_worksheet.cell(row=curr_row, column=3).value = total_amount
        output_worksheet.cell(row=curr_row, column=3).font = Font(bold=True)
        output_worksheet.cell(row=curr_row, column=9).value = total_commission
        output_worksheet.cell(row=curr_row, column=9).font = Font(bold=True)
        curr_row += 1


    output_wb.save(filename = dest_filename)

def run(e):
    try:
        global root
        seen = set()
        bills = [bill.upper() for bill in e['bills'].get("1.0","end-1c").split('\n') if bill and bill not in seen and not seen.add(bill)]

        dest_filename='output.xlsx'
        rate_engine = float(e['เครื่องยนต์'].get())
        rate_part = float(e['อะไหล่'].get())

        # bills = scan_bill()
        output = get_data(bills,server="localhost",database='db',username='uid',password = 'password',rate_engine=rate_engine,rate_part=rate_part)
        write_output(output,dest_filename=dest_filename)
        
        os.system(f'start excel {dest_filename}')

        root.destroy()

    except ValueError:
        tkinter.messagebox.showerror("Error", "เกิดข้อผิดพลาด")
    except PermissionError:
        tkinter.messagebox.showerror("Error", "โปรดปิดไฟล์เก่าก่อน")

def open_file(e):
        global extra_commission
        extra_commission = {}
        fname = filedialog.askopenfilename(filetypes=(("Text files", "*.txt"),("All files", "*.*") ))
        try:
            with open(fname) as file:
                e['display'].config(state='normal')
                e['display'].delete('1.0', tkinter.END)
                for line in file.readlines():
                    if line.strip():
                        item_id,percent = line.strip().split(' ')
                        extra_commission[item_id] = float(percent)
                        e['display'].insert(tkinter.INSERT,item_id + '\t' +percent + '\n')
                e['display'].config(state='disabled')
       
        except FileExistsError:                    
            tkinter.messagebox.showerror("Error", "Failed to read file\n'%s'" % fname)
        except ValueError:
            tkinter.messagebox.showerror("Error", "ไฟล์ไม่ถูกต้อง")
            

def makeform(root, fields=['เครื่องยนต์','อะไหล่']):
    entries = {}
    row = tkinter.Frame(root)
    row.pack(side=tkinter.TOP, fill=tkinter.X, padx=5)
    lab = tkinter.Label(row, text='ใส่เลขที่บิล', anchor='w',font=(20))
    lab.pack(side=tkinter.TOP)
    
    row = tkinter.Frame(root)
    row.pack(side=tkinter.TOP, fill=tkinter.X, padx=5)
    editor = EditorClass(row)
    entries['bills'] = editor.text
  
    for field in fields:
        row = tkinter.Frame(root)
        lab = tkinter.Label(row, width=10, text=field, anchor='w',font=(20))
        ent = tkinter.Entry(row)
        row.pack(side=tkinter.TOP, fill=tkinter.X, padx=5, pady=5)
        lab.pack(side=tkinter.LEFT)
        ent.pack(side=tkinter.RIGHT, expand=tkinter.YES, fill=tkinter.X)
        entries[field] = ent

    row = tkinter.Frame(root)
    configfile = tkinter.Text(row,width=40, height='10',state='disabled',background = 'lightgrey')
    configfile.pack(side=tkinter.TOP)
    row.pack(side=tkinter.TOP, fill=tkinter.X, padx=5, pady=5)
    entries['display'] = configfile

    return entries


def main():
    global root
    
    root.title("Commission")
    root.minsize(100,100)

    ents = makeform(root)

    ok_btn = tkinter.Button(root,text='OK',font=(20),command=(lambda e=ents: run(e)))
    ok_btn.pack(side=tkinter.TOP, padx=5, pady=5)


    open_btn = tkinter.Button(root,text='เปิดไฟล์',font=(20),command=(lambda e=ents: open_file(e)))
    open_btn.pack(side=tkinter.TOP, padx=5, pady=5)

    root.mainloop()
    

if __name__ == '__main__':
    main()

# '''
# gid
# 1000 เครื่องยน
# 1001 อะไหล่
# 1002 ทั่วไป
# 2000 ตลับเมตร
# 4000 สมาคุณ
# 4001 มัเทำครัว
# '''
